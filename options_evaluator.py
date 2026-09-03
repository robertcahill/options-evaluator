"""
================================================================================
OPTIONS STRATEGY EVALUATOR  —  Phase 1
================================================================================
A learning-oriented tool for evaluating option strategies before you trade them.

WHAT THIS DOES
    You describe a strategy (its legs), give it the current stock price and each
    option's implied volatility, and it computes:
        - the Greeks for each leg AND the net Greeks of the whole position
        - max profit, max loss, and breakeven(s) from the STRUCTURE, not a chart
        - probability of profit from breakevens + a lognormal at expiry
        - Black-Scholes value vs the mid you typed (and IV back-solved from mid)
        - capital tied up and return-on-capital
        - a payoff diagram (expiration + T+0) with a 1-sigma expected-move band
    ...then writes a clean Excel dashboard and shows a payoff chart.

WHY IT'S BUILT THIS WAY (for your Python learning)
    The code is organized in layers, bottom to top:
        1. The math engine   (Black-Scholes + Greeks)   <- pure functions
        2. The building block (an OptionLeg)             <- one contract
        3. The strategy       (a list of legs + stock)   <- the position
        4. The strategies you can build  (covered call, CSP, long call)
        5. The outputs        (console summary, Excel, chart)
    Each layer only uses the layer below it. Once you see how ONE strategy is
    assembled from legs, adding a new one (iron condor, straddle, etc.) is the
    same pattern repeated. That extensibility is deliberate — see the note at the
    very bottom.

KEY IDEA TO CARRY (the hinge for everything):
    Implied Volatility (IV) is what you're really negotiating over.
      - BUYING premium (long call/put) wants LOW IV  — options are cheap.
      - SELLING premium (covered call, cash-secured put) wants HIGH IV — you
        collect more for the same risk.
    The same tool serves both sides because the core question — "is vol rich or
    cheap?" — drives buy vs. sell in opposite directions.

    Use the CONTRACT'S IV from the chain row, not the blended "Implied Volatility"
    on Today's Options Statistics. If you mix those, Greeks and the model price
    will not match the quote.
================================================================================
"""

import math
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional

import numpy as np
from scipy.stats import norm   # norm.cdf / norm.pdf = the normal distribution
                               # functions Black-Scholes needs


# =============================================================================
# LAYER 1 — THE MATH ENGINE (Black-Scholes pricing + the Greeks)
# =============================================================================
# These are pure functions: same inputs always give same outputs, no side
# effects. That makes them easy to test and easy to trust. Every options tool
# on earth is built on top of these five-ish formulas.
#
# The six inputs Black-Scholes needs for a European option:
#   S     = current price of the underlying stock
#   K     = strike price of the option
#   T     = time to expiration, IN YEARS (e.g. 30 days = 30/365 = 0.082)
#   r     = risk-free interest rate (annual, decimal — e.g. 0.045 for 4.5%)
#   sigma = implied volatility (annual, decimal — e.g. 0.60 for 60%)
#   q     = dividend yield (annual, decimal — 0 if you don't want to model it)
# =============================================================================

def _d1_d2(S, K, T, r, sigma, q=0.0):
    """
    d1 and d2 are the two intermediate terms that show up all over the
    Black-Scholes formulas. We compute them once here and reuse them, rather
    than repeating the algebra in every function below.

    Intuition (rough): d1/d2 measure how many standard deviations the option is
    'in the money' in probability space. You don't need to memorize the formula;
    you need to know these feed everything else.
    """
    # Guard against T=0 (expiration) which would divide by zero.
    if T <= 0:
        T = 1e-9
    d1 = (np.log(S / K) + (r - q + 0.5 * sigma ** 2) * T) / (sigma * np.sqrt(T))
    d2 = d1 - sigma * np.sqrt(T)
    return d1, d2


def bs_price(S, K, T, r, sigma, kind, q=0.0):
    """
    The Black-Scholes fair price of a single European option.
    kind = 'call' or 'put'.

    This is the theoretical value at the IV you typed. Comparing it to the
    ACTUAL mid is a SANITY CHECK more than an alpha signal: if you pulled IV
    and mid from the same chain row, they should nearly agree. A big gap means
    mixed IV sources, a bad quote, or American/dividend effects this model
    doesn't capture.
    """
    d1, d2 = _d1_d2(S, K, T, r, sigma, q)
    df_q = np.exp(-q * T)
    df_r = np.exp(-r * T)
    if kind == 'call':
        return S * df_q * norm.cdf(d1) - K * df_r * norm.cdf(d2)
    else:  # put
        return K * df_r * norm.cdf(-d2) - S * df_q * norm.cdf(-d1)


def implied_vol(S, K, T, r, market_price, kind, q=0.0):
    """
    Back-solve the IV that makes Black-Scholes equal the mid you typed.

    Why this exists: you type TWO vol-related numbers (IV and premium). This
    is how we check they belong to the same contract. Returns None if the
    premium is at/below European intrinsic (can't invert).
    """
    if market_price is None or market_price <= 0 or T <= 0 or S <= 0 or K <= 0:
        return None
    df_r = math.exp(-r * T)
    df_q = math.exp(-q * T)
    if kind == 'call':
        floor = max(S * df_q - K * df_r, 0.0)
    else:
        floor = max(K * df_r - S * df_q, 0.0)
    # A mid at or below discounted intrinsic cannot be a European IV.
    if market_price <= floor + 1e-8:
        return None

    lo, hi = 1e-4, 5.0
    for _ in range(80):
        mid_sig = 0.5 * (lo + hi)
        model = bs_price(S, K, T, r, mid_sig, kind, q)
        if model > market_price:
            hi = mid_sig
        else:
            lo = mid_sig
    return 0.5 * (lo + hi)


def prob_spot_above(S, barrier, T, r, sigma, q=0.0):
    """
    Risk-neutral P(S_T > barrier) under Black-Scholes. This is N(d2) with
    K = barrier — the probability the stock finishes above a price, NOT delta
    (delta is N(d1) and includes a hedge-ratio extra).
    """
    if barrier <= 0:
        return 1.0
    if T <= 0:
        return 1.0 if S > barrier else 0.0
    if sigma <= 0:
        fwd = S * math.exp((r - q) * T)
        return 1.0 if fwd > barrier else 0.0
    _, d2 = _d1_d2(S, barrier, T, r, sigma, q)
    return float(norm.cdf(d2))


# ---- THE GREEKS -------------------------------------------------------------
# Each Greek is the sensitivity of the option PRICE to one input, holding the
# others fixed. This is the decomposition of "what am I paying for / exposed to."

def greek_delta(S, K, T, r, sigma, kind, q=0.0):
    """
    DELTA = sensitivity to a $1 move in the stock (first-order direction).
      - Ranges 0..1 for calls, -1..0 for puts (times e^{-qT}).
      - NOT the same as probability of finishing ITM — that's N(d2). We use
        d2 for probability-of-profit and keep delta for directional exposure.
    """
    d1, _ = _d1_d2(S, K, T, r, sigma, q)
    disc = np.exp(-q * T)
    if kind == 'call':
        return disc * norm.cdf(d1)
    else:
        return disc * (norm.cdf(d1) - 1.0)


def greek_gamma(S, K, T, r, sigma, q=0.0):
    """
    GAMMA = rate of change of delta (convexity). Same for calls and puts.
      - Highest for at-the-money, near-expiry options.
      - This is the 'acceleration' you're long when you BUY options (and short
        when you sell them). Long gamma = your winners speed up. You pay for it
        in theta.
    """
    d1, _ = _d1_d2(S, K, T, r, sigma, q)
    t = T if T > 0 else 1e-9
    return np.exp(-q * T) * norm.pdf(d1) / (S * sigma * np.sqrt(t))


def greek_theta(S, K, T, r, sigma, kind, q=0.0):
    """
    THETA = time decay per DAY (we divide the annual figure by 365).
      - Negative when you're LONG an option (value bleeds away each day) — this
        is the 'rent' you pay for gamma.
      - POSITIVE when you're SHORT an option (you collect decay) — this is the
        engine of income-selling strategies.
    """
    d1, d2 = _d1_d2(S, K, T, r, sigma, q)
    t = T if T > 0 else 1e-9
    df_q = np.exp(-q * T)
    df_r = np.exp(-r * T)
    first = -(S * df_q * norm.pdf(d1) * sigma) / (2 * np.sqrt(t))
    if kind == 'call':
        annual_theta = (first
                        - r * K * df_r * norm.cdf(d2)
                        + q * S * df_q * norm.cdf(d1))
    else:
        annual_theta = (first
                        + r * K * df_r * norm.cdf(-d2)
                        - q * S * df_q * norm.cdf(-d1))
    return annual_theta / 365.0


def greek_vega(S, K, T, r, sigma, q=0.0):
    """
    VEGA = sensitivity to a 1-POINT (1%) change in implied volatility. Same for
    calls and puts. (We divide by 100 so it reads 'per 1% IV move'.)
      - Long options = long vega. If IV falls after you buy, you lose on vega
        even if direction was right.
      - Short options = short vega = you WANT IV to fall (IV crush after earnings
        is a premium-seller's friend).
    """
    d1, _ = _d1_d2(S, K, T, r, sigma, q)
    t = T if T > 0 else 1e-9
    return S * np.exp(-q * T) * norm.pdf(d1) * np.sqrt(t) / 100.0


def greek_rho(S, K, T, r, sigma, kind, q=0.0):
    """
    RHO = sensitivity to a 1% change in interest rates. Usually the least
    important Greek — but matters more for LONG-DATED options (LEAPs) and when
    rates are moving. Divided by 100 to read 'per 1% rate move.'
    """
    _, d2 = _d1_d2(S, K, T, r, sigma, q)
    if kind == 'call':
        return K * T * np.exp(-r * T) * norm.cdf(d2) / 100.0
    else:
        return -K * T * np.exp(-r * T) * norm.cdf(-d2) / 100.0


# =============================================================================
# LAYER 2 — THE BUILDING BLOCK: one OptionLeg
# =============================================================================
# Every strategy, no matter how complex, is just a COMBINATION OF LEGS. A leg is
# a single option contract (or a stock position) that you're either long or
# short. Get this one class right and every strategy is just a list of these.
#
# @dataclass is a Python shortcut: it auto-writes the boilerplate __init__ for
# you. You just declare the fields and their types. Great for 'bags of data'.
# =============================================================================

@dataclass
class OptionLeg:
    kind: str          # 'call', 'put', or 'stock'
    strike: float      # the strike price (ignored for stock legs)
    iv: float          # implied volatility for THIS contract, as a decimal (0.60 = 60%)
    quantity: int      # +1 = long one contract, -1 = short one contract
    premium: float = 0.0   # what you paid (long) or received (short) per share
    # note: 1 option contract = 100 shares, handled when we sum dollars later

    def value_at_expiration(self, stock_price):
        """
        The INTRINSIC value of this leg at expiration, per share, for a given
        final stock price. This is what the leg is worth when all time value is
        gone — the foundation of the payoff diagram.
        """
        if self.kind == 'call':
            intrinsic = max(stock_price - self.strike, 0.0)
        elif self.kind == 'put':
            intrinsic = max(self.strike - stock_price, 0.0)
        else:  # stock leg — worth the stock price itself
            intrinsic = stock_price
        return intrinsic * self.quantity

    def cost_basis(self):
        """
        The cash flow to OPEN this leg, per share.
        Long (quantity +1): you PAY the premium  -> negative cash (money out).
        Short (quantity -1): you RECEIVE premium  -> positive cash (money in).
        For a stock leg, 'premium' holds the entry price.
        """
        return -self.premium * self.quantity


# =============================================================================
# LAYER 3 — THE STRATEGY: a list of legs + the current stock price
# =============================================================================
# This is where the useful analysis lives. A Strategy holds the legs and the
# market context (stock price, rate, days to expiry) and answers the questions
# you actually care about: net Greeks, max profit/loss, breakevens, payoff.
# =============================================================================

@dataclass
class Strategy:
    name: str
    legs: List[OptionLeg]
    stock_price: float        # current price of the underlying (S)
    days_to_expiry: int       # calendar days until the options expire
    risk_free_rate: float = 0.045   # r — a reasonable default; update to current
    dividend_yield: float = 0.0     # q — 0 unless you want to model dividends
    iv_percentile: Optional[float] = None  # 0-100; optional context, not used in math

    @property
    def T(self):
        """Time to expiry in YEARS — what the math engine needs."""
        return self.days_to_expiry / 365.0

    @property
    def q(self):
        return self.dividend_yield

    def option_legs(self):
        return [leg for leg in self.legs if leg.kind in ('call', 'put')]

    def representative_iv(self):
        """Average of the option-leg IVs. Used for expected-move and PoP."""
        ivs = [leg.iv for leg in self.option_legs() if leg.iv and leg.iv > 0]
        if not ivs:
            return None
        return float(np.mean(ivs))

    # ---- NET GREEKS ---------------------------------------------------------
    # We sum each Greek across all legs, weighting by quantity (long/short) and
    # by 100 (shares per contract) so the numbers are in DOLLARS of P&L per unit
    # move. Net Greeks tell you the whole position's exposure at a glance:
    #   net delta  -> your directional bias right now
    #   net theta  -> are you PAYING or COLLECTING time decay each day
    #   net vega   -> do you WIN or LOSE if IV rises
    def _leg_greeks(self, leg):
        """Dollar Greeks for one leg (already signed and x100)."""
        g = {'delta': 0.0, 'gamma': 0.0, 'theta': 0.0, 'vega': 0.0, 'rho': 0.0}
        if leg.kind == 'stock':
            g['delta'] = 1.0 * leg.quantity * 100
            return g
        S, K, T, r, sig, q = (self.stock_price, leg.strike, self.T,
                              self.risk_free_rate, leg.iv, self.q)
        mult = leg.quantity * 100
        g['delta'] = greek_delta(S, K, T, r, sig, leg.kind, q) * mult
        g['gamma'] = greek_gamma(S, K, T, r, sig, q) * mult
        g['theta'] = greek_theta(S, K, T, r, sig, leg.kind, q) * mult
        g['vega']  = greek_vega(S, K, T, r, sig, q) * mult
        g['rho']   = greek_rho(S, K, T, r, sig, leg.kind, q) * mult
        return g

    def net_greeks(self):
        totals = {'delta': 0.0, 'gamma': 0.0, 'theta': 0.0, 'vega': 0.0, 'rho': 0.0}
        for leg in self.legs:
            for k, v in self._leg_greeks(leg).items():
                totals[k] += v
        return totals

    def per_leg_detail(self):
        """Per-leg Greeks plus model price vs the mid you typed."""
        rows = []
        for leg in self.legs:
            row = {
                'leg': leg,
                'greeks': self._leg_greeks(leg),
                'model_price': None,
                'implied_vol': None,
            }
            if leg.kind in ('call', 'put'):
                row['model_price'] = float(bs_price(
                    self.stock_price, leg.strike, self.T, self.risk_free_rate,
                    leg.iv, leg.kind, self.q))
                row['implied_vol'] = implied_vol(
                    self.stock_price, leg.strike, self.T, self.risk_free_rate,
                    leg.premium, leg.kind, self.q)
            rows.append(row)
        return rows

    # ---- PAYOFF AT EXPIRATION ----------------------------------------------
    # For a given final stock price, total profit/loss (in dollars) =
    #   (what all legs are worth at expiration)  +  (the cash you took in/paid
    #    to open them)  ... all x100 for shares-per-contract.
    def profit_at(self, stock_price):
        total = 0.0
        for leg in self.legs:
            total += leg.value_at_expiration(stock_price) * 100   # value at expiry
            total += leg.cost_basis() * 100                       # opening cash flow
        return total

    def mark_pnl(self, stock_price, days_left=None):
        """
        Model P&L if spot is `stock_price` and `days_left` remain (T+0 when
        days_left is the original DTE). Uses Black-Scholes at each leg's IV.
        At days_left=0 this collapses toward expiration intrinsic.
        """
        if days_left is None:
            days_left = self.days_to_expiry
        T = max(days_left, 0) / 365.0
        total = 0.0
        for leg in self.legs:
            total += leg.cost_basis() * 100
            if leg.kind == 'stock':
                total += stock_price * leg.quantity * 100
            else:
                px = bs_price(stock_price, leg.strike, T, self.risk_free_rate,
                              leg.iv, leg.kind, self.q)
                total += px * leg.quantity * 100
        return total

    def payoff_curve(self, low=None, high=None, points=200):
        """
        Build arrays of (stock_price, profit) across a range, for the diagram.
        Defaults to +/- 40% around the current price if no range given.
        This WINDOW is for the picture only — max profit/loss do NOT come from it.
        """
        if low is None:
            low = self.stock_price * 0.6
        if high is None:
            high = self.stock_price * 1.4
        prices = np.linspace(low, high, points)
        profits = np.array([self.profit_at(p) for p in prices])
        return prices, profits

    def expected_move(self):
        """
        Trader 1-sigma expected move through expiry: S × IV × √T.
        This is a convention, not a precise lognormal band. Drawn on the chart
        so the wings are not mistaken for 'likely' outcomes.
        """
        iv = self.representative_iv()
        if iv is None:
            return None
        return self.stock_price * iv * math.sqrt(max(self.T, 0.0))

    # ---- STRUCTURED P/L (not the chart window) ------------------------------
    def _expiration_slope(self, spot):
        """Dollars of expiration P&L per $1 of stock, just above `spot`."""
        x = spot + 1e-9
        slope = 0.0
        for leg in self.legs:
            if leg.kind == 'stock':
                slope += 100.0 * leg.quantity
            elif leg.kind == 'call' and x > leg.strike:
                slope += 100.0 * leg.quantity
            elif leg.kind == 'put' and x < leg.strike:
                slope += -100.0 * leg.quantity
        return slope

    def _knots(self):
        """Prices where expiration P&L kinks (0, every strike, a far-right point)."""
        strikes = [leg.strike for leg in self.option_legs()]
        far = max([self.stock_price, *strikes] or [self.stock_price]) * 3.0
        return sorted(set([0.0] + strikes + [far]))

    def structural_extrema(self):
        """
        True max profit / max loss at expiration from the payoff's shape.

        Expiration P&L is piecewise-linear, with kinks only at strikes. So the
        finite extrema live at 0 and at the strikes. If the slope as S→∞ is
        positive (naked long call / long stock), max profit is unlimited; if
        negative (naked short call), max loss is unlimited.
        """
        values = [self.profit_at(p) for p in self._knots()]
        max_profit = max(values)
        max_loss = min(values)
        slope_inf = self._expiration_slope(1e12)
        if slope_inf > 1e-6:
            max_profit = math.inf
        if slope_inf < -1e-6:
            max_loss = -math.inf
        return max_profit, max_loss

    def breakevens(self):
        """
        Exact expiration breakevens: P&L is linear between knots, so a sign
        change between two adjacent knots interpolates to the true zero.
        """
        knots = self._knots()
        found = []
        for i in range(1, len(knots)):
            x0, x1 = knots[i - 1], knots[i]
            y0, y1 = self.profit_at(x0), self.profit_at(x1)
            if abs(y0) < 1e-8:
                found.append(x0)
            elif y0 * y1 < 0:
                found.append(x0 - y0 * (x1 - x0) / (y1 - y0))
        if abs(self.profit_at(knots[-1])) < 1e-8:
            found.append(knots[-1])
        # Deduplicate (can hit a knot exactly from both sides).
        cleaned = []
        for b in found:
            if b < -1e-6:
                continue
            if not cleaned or abs(b - cleaned[-1]) > 1e-4:
                cleaned.append(b)
        return cleaned

    def net_cash_to_open(self):
        """Negative = you paid (debit); positive = you collected (credit)."""
        return sum(leg.cost_basis() * 100 for leg in self.legs)

    def model_edge_dollars(self):
        """
        Sum over option legs of (model − mid) × quantity × 100.

        Positive means Black-Scholes at YOUR typed IVs says the position is
        cheaper to buy / richer to sell than the mids. If you typed each
        contract's own IV, this should be near zero — a large number is a
        warning, not a free lunch.
        """
        edge = 0.0
        any_opt = False
        for row in self.per_leg_detail():
            if row['model_price'] is None:
                continue
            any_opt = True
            edge += (row['model_price'] - row['leg'].premium) * row['leg'].quantity * 100
        return edge if any_opt else None

    def capital_required(self):
        """
        Conservative cash to put the trade on. Heuristic, not your broker's
        margin:
          - long stock (covered call): pay for shares, net of option credit/debit
          - naked short put (CSP): cash-secure at strike × 100
          - long premium: the debit
          - naked short call: None (undefined here — that's margin)
        """
        stock_legs = [leg for leg in self.legs if leg.kind == 'stock']
        short_calls = [leg for leg in self.legs if leg.kind == 'call' and leg.quantity < 0]
        long_calls = [leg for leg in self.legs if leg.kind == 'call' and leg.quantity > 0]
        short_puts = [leg for leg in self.legs if leg.kind == 'put' and leg.quantity < 0]
        long_puts = [leg for leg in self.legs if leg.kind == 'put' and leg.quantity > 0]
        long_stock_qty = sum(leg.quantity for leg in stock_legs if leg.quantity > 0)

        # Naked short call (no stock, no long call to define the risk) → margin.
        if short_calls and long_stock_qty <= 0 and not long_calls:
            return None

        if long_stock_qty > 0:
            stock_outlay = sum(leg.premium * leg.quantity * 100
                               for leg in stock_legs if leg.quantity > 0)
            opt_cf = sum(leg.cost_basis() * 100 for leg in self.option_legs())
            return stock_outlay - opt_cf

        # Cash-secured naked put: set aside strike × 100.
        if short_puts and not long_puts:
            return sum(leg.strike * 100 * abs(leg.quantity) for leg in short_puts)

        # Defined-risk verticals: width, minus credit if you collected one.
        if short_puts and long_puts:
            width = 0.0
            for s in short_puts:
                lowers = [lp for lp in long_puts if lp.strike <= s.strike]
                if lowers:
                    k_long = max(lp.strike for lp in lowers)
                    width += abs(s.strike - k_long) * 100 * abs(s.quantity)
            credit = max(self.net_cash_to_open(), 0.0)
            return max(width - credit, 0.0)
        if short_calls and long_calls and long_stock_qty <= 0:
            width = 0.0
            for s in short_calls:
                highers = [lc for lc in long_calls if lc.strike >= s.strike]
                if highers:
                    k_long = min(lc.strike for lc in highers)
                    width += abs(k_long - s.strike) * 100 * abs(s.quantity)
            if width > 0:
                credit = max(self.net_cash_to_open(), 0.0)
                if credit > 0:
                    return max(width - credit, 0.0)
                return max(-self.net_cash_to_open(), 0.0)

        debit = -self.net_cash_to_open()
        return debit if debit > 0 else abs(self.net_cash_to_open())

    def summary(self):
        max_profit, max_loss = self.structural_extrema()
        capital = self.capital_required()
        net = self.net_cash_to_open()
        roc = None
        if capital and capital > 0 and math.isfinite(max_profit):
            roc = max_profit / capital
        return {
            'max_profit': max_profit,
            'max_loss': max_loss,
            'breakevens': self.breakevens(),
            'net_cost_or_credit': net,
            'capital': capital,
            'roc_at_max_profit': roc,
            'model_edge': self.model_edge_dollars(),
        }

    # ---- PROBABILITY OF PROFIT ----------------------------------------------
    # Expiration P&L is profitable on one or more intervals of the stock-price
    # line, split by breakevens. We put a lognormal at expiry (using the
    # average of the option IVs) and add P(S_T in each profitable interval).
    #
    # This is risk-neutral PoP — the same world Black-Scholes lives in — not
    # a forecast with a real-world drift. Better than 1−|delta|, still not gospel.
    def probability_of_profit_estimate(self):
        sigma = self.representative_iv()
        if sigma is None or sigma <= 0:
            return None
        bes = self.breakevens()
        edges = [0.0] + bes + [None]  # None = +∞
        pop = 0.0
        for lo, hi in zip(edges, edges[1:]):
            probe = (lo + hi) / 2.0 if hi is not None else max(lo * 1.05, lo + 1.0, 1.0)
            if hi is not None and hi <= lo:
                continue
            if self.profit_at(probe) <= 0:
                continue
            p_above_lo = prob_spot_above(self.stock_price, lo, self.T,
                                         self.risk_free_rate, sigma, self.q)
            p_above_hi = (0.0 if hi is None else
                          prob_spot_above(self.stock_price, hi, self.T,
                                          self.risk_free_rate, sigma, self.q))
            pop += max(p_above_lo - p_above_hi, 0.0)
        return min(max(pop, 0.0), 1.0)


# =============================================================================
# LAYER 4 — THE STRATEGIES YOU CAN BUILD
# =============================================================================
# Each function here is just a convenient way to assemble the right legs. Notice
# the PATTERN: every strategy is "make some OptionLegs, hand them to Strategy()."
# Once you see these three, you can add any strategy by copying the pattern.
#
# We include:
#   1. long_call        — a directional/convexity BUY (your Tesla/Marvell trades)
#   2. covered_call     — an INCOME strategy (own stock, sell a call against it)
#   3. cash_secured_put — an INCOME strategy (sell a put, ready to buy the stock)
# The two income strategies are where you'd start for "selling for income."
# =============================================================================

def _strategy_kwargs(r, q, iv_percentile):
    return dict(risk_free_rate=r, dividend_yield=q, iv_percentile=iv_percentile)


def long_call(stock_price, strike, iv, premium, days_to_expiry,
              r=0.045, q=0.0, iv_percentile=None):
    """
    LONG CALL — you PAY premium for the right to buy at 'strike'.
    Risk: limited to the premium. Reward: unlimited. You are LONG gamma & vega,
    SHORT theta (paying rent). Wants LOW IV at entry (cheap) + a move up.
    """
    legs = [OptionLeg(kind='call', strike=strike, iv=iv, quantity=+1, premium=premium)]
    return Strategy(f"Long {strike}C", legs, stock_price, days_to_expiry,
                    **_strategy_kwargs(r, q, iv_percentile))


def covered_call(stock_price, stock_entry, call_strike, call_iv, call_premium,
                 days_to_expiry, r=0.045, q=0.0, iv_percentile=None):
    """
    COVERED CALL — you OWN 100 shares and SELL a call against them for income.
    You collect the call premium (positive theta — decay works FOR you).
    Trade-off: your upside is CAPPED at the strike (if the stock rockets past it,
    your shares get called away and you miss the rest). Downside: you still own
    the stock, so a big drop hurts — the premium only cushions it.
    Best when IV is HIGH (fat premium) and you're neutral-to-mildly-bullish.
    """
    legs = [
        OptionLeg(kind='stock', strike=0, iv=0, quantity=+1, premium=stock_entry),
        OptionLeg(kind='call', strike=call_strike, iv=call_iv, quantity=-1,
                  premium=call_premium),   # quantity -1 = SHORT (we sold it)
    ]
    return Strategy(f"Covered Call {call_strike}C", legs, stock_price, days_to_expiry,
                    **_strategy_kwargs(r, q, iv_percentile))


def cash_secured_put(stock_price, put_strike, put_iv, put_premium,
                     days_to_expiry, r=0.045, q=0.0, iv_percentile=None):
    """
    CASH-SECURED PUT — you SELL a put and set aside cash to buy the stock if
    assigned. You collect the put premium (positive theta). If the stock stays
    above the strike, you keep the premium as income. If it falls below, you're
    OBLIGATED to buy 100 shares at the strike (your max loss is large: strike
    minus premium, all the way down to zero).
    This is the 'get paid to place a limit buy order' income trade. Best when IV
    is HIGH and you'd genuinely be happy owning the stock at the strike.
    """
    legs = [OptionLeg(kind='put', strike=put_strike, iv=put_iv, quantity=-1,
                      premium=put_premium)]
    return Strategy(f"Cash-Secured Put {put_strike}P", legs, stock_price, days_to_expiry,
                    **_strategy_kwargs(r, q, iv_percentile))


# =============================================================================
# LAYER 5 — THE OUTPUTS (console summary, payoff chart, Excel dashboard)
# =============================================================================

def _fmt_money(x):
    if x is None:
        return "n/a"
    if x == math.inf or x == -math.inf:
        return "unlimited"
    return f"${x:,.2f}"


def _iv_bias_line(percentile):
    if percentile is None:
        return None
    if percentile >= 70:
        return (f"IV percentile {percentile:.0f} -> vol is RICH; "
                "selling premium is the usual lean.")
    if percentile <= 30:
        return (f"IV percentile {percentile:.0f} -> vol is CHEAP; "
                "buying premium is the usual lean.")
    return (f"IV percentile {percentile:.0f} -> vol is mid-range; "
            "let the structure (not percentile) decide.")


def print_summary(strat: Strategy):
    """Readable console report — the fast 'what am I looking at' view."""
    s = strat.summary()
    g = strat.net_greeks()
    pop = strat.probability_of_profit_estimate()
    details = strat.per_leg_detail()

    print("=" * 64)
    print(f"  STRATEGY: {strat.name}")
    print(f"  Stock price: ${strat.stock_price:.2f}   Days to expiry: {strat.days_to_expiry}")
    if strat.dividend_yield:
        print(f"  Dividend yield: {strat.dividend_yield*100:.2f}%")
    print("=" * 64)

    credit_debit = ("CREDIT (you collect)" if s['net_cost_or_credit'] > 0
                    else "DEBIT (you pay)")
    print(f"  Net to open:      {_fmt_money(abs(s['net_cost_or_credit']))}  [{credit_debit}]")
    print(f"  Max profit:       {_fmt_money(s['max_profit'])}")
    print(f"  Max loss:         {_fmt_money(s['max_loss'])}")
    be_str = ", ".join(f"${b:.2f}" for b in s['breakevens']) or "n/a"
    print(f"  Breakeven(s):     {be_str}")
    if s['capital'] is None:
        print("  Capital:          n/a (naked short call - broker margin)")
    else:
        print(f"  Capital:          {_fmt_money(s['capital'])}")
    if s['roc_at_max_profit'] is not None:
        print(f"  ROC at max:       {s['roc_at_max_profit']*100:.1f}%  "
              "(max profit / capital)")
    if pop is not None:
        print(f"  Prob. of profit:  ~{pop*100:.0f}%  "
              "(risk-neutral, from breakevens)")
    bias = _iv_bias_line(strat.iv_percentile)
    if bias:
        print(f"  {bias}")
    print("-" * 64)
    print("  NET GREEKS (dollar P&L per unit move, whole position):")
    print(f"    Delta: {g['delta']:+.1f}   (directional bias; + = bullish)")
    print(f"    Gamma: {g['gamma']:+.2f}   (convexity; + = you own acceleration)")
    print(f"    Theta: {g['theta']:+.2f}   (per DAY; + = you COLLECT decay = income)")
    print(f"    Vega:  {g['vega']:+.2f}   (per 1% IV; + = you gain if IV rises)")
    print(f"    Rho:   {g['rho']:+.2f}")
    print("-" * 64)
    print("  LEGS  (model vs mid is a consistency check, not a mispricing signal)")
    print(f"    {'Type':<8} {'K':>8} {'Qty':>5} {'Mid':>8} {'Model':>8} "
          f"{'IVtyped':>8} {'IVmid':>8} {'Delta$':>9}")
    for row in details:
        leg = row['leg']
        k = f"{leg.strike:.1f}" if leg.kind != 'stock' else "-"
        model = f"{row['model_price']:.2f}" if row['model_price'] is not None else "-"
        iv_t = f"{leg.iv*100:.1f}%" if leg.kind != 'stock' else "-"
        iv_m = (f"{row['implied_vol']*100:.1f}%" if row['implied_vol'] is not None
                else ("n/a" if leg.kind != 'stock' else "-"))
        print(f"    {leg.kind:<8} {k:>8} {leg.quantity:>+5d} {leg.premium:>8.2f} "
              f"{model:>8} {iv_t:>8} {iv_m:>8} {row['greeks']['delta']:>+9.1f}")
        if (row['implied_vol'] is not None and
                abs(row['implied_vol'] - leg.iv) > 0.03):
            print("      ^ typed IV and mid-implied IV differ by >3 vol points. "
                  "Check you used the CONTRACT row's IV.")
    if s['model_edge'] is not None:
        print(f"    Model edge vs mids: {_fmt_money(s['model_edge'])}  "
              "(~0 if IV and mid are from the same row)")
    print("-" * 64)
    if g['theta'] > 0:
        print("  READ: Positive theta -> time decay works FOR you (income profile).")
    else:
        print("  READ: Negative theta -> you're paying time decay (long-premium).")
    if g['vega'] < 0:
        print("        Negative vega  -> you WANT IV to fall (good after a vol spike).")
    else:
        print("        Positive vega  -> you gain if IV rises (careful buying high IV).")
    print("=" * 64)
    print()


def make_payoff_chart(strat: Strategy, path):
    """Save the payoff diagram: expiration, T+0 (today's mark), expected-move band."""
    import matplotlib
    matplotlib.use('Agg')   # non-interactive backend (we're saving to file)
    import matplotlib.pyplot as plt

    prices, profits = strat.payoff_curve()
    t0 = np.array([strat.mark_pnl(p) for p in prices])
    em = strat.expected_move()

    fig, ax = plt.subplots(figsize=(9, 5))
    ax.plot(prices, profits, linewidth=2, color='#1f4e79',
            label="At expiration")
    ax.plot(prices, t0, linewidth=1.5, color='#1f4e79', linestyle='--',
            label="T+0 (model, IV held constant)")
    ax.axhline(0, color='#888', linewidth=1)
    ax.axvline(strat.stock_price, color='#c00', linestyle='--',
               linewidth=1, label=f"Current ${strat.stock_price:.0f}")
    if em:
        ax.axvspan(strat.stock_price - em, strat.stock_price + em,
                   color='#1f4e79', alpha=0.08,
                   label=f"1σ expected move ±${em:.0f}")
    ax.fill_between(prices, profits, 0, where=(profits >= 0),
                    color='#2e7d32', alpha=0.12)
    ax.fill_between(prices, profits, 0, where=(profits < 0),
                    color='#c62828', alpha=0.12)
    ax.set_title(f"Payoff — {strat.name}")
    ax.set_xlabel("Stock price ($)")
    ax.set_ylabel("Profit / Loss ($)")
    ax.legend(loc="best", fontsize=8)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(path, dpi=120)
    plt.close(fig)


def export_excel(strat: Strategy, path):
    """Write a clean Excel dashboard: summary stats, per-leg detail, net Greeks,
    and the payoff table. This is the workflow artifact you keep/share."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.chart import LineChart, Reference

    wb = Workbook()
    ws = wb.active
    ws.title = "Strategy"

    header = Font(bold=True, size=13, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="1F4E79")
    bold = Font(bold=True)

    def title_row(r, text):
        ws.cell(r, 1, text).font = header
        ws.cell(r, 1).fill = hfill
        for c in range(2, 9):
            ws.cell(r, c).fill = hfill

    def money_cell(value):
        if value is None:
            return "n/a"
        if value == math.inf or value == -math.inf:
            return "unlimited"
        return round(float(value), 2)

    s = strat.summary()
    g = strat.net_greeks()
    pop = strat.probability_of_profit_estimate()

    title_row(1, f"OPTIONS STRATEGY:  {strat.name}")
    ws.cell(2, 1, "Stock price"); ws.cell(2, 2, round(strat.stock_price, 2))
    ws.cell(3, 1, "Days to expiry"); ws.cell(3, 2, strat.days_to_expiry)
    ws.cell(4, 1, "Net to open ($)"); ws.cell(4, 2, money_cell(s['net_cost_or_credit']))
    ws.cell(4, 3, "(+ = credit/collect, - = debit/pay)")
    ws.cell(5, 1, "Max profit ($)"); ws.cell(5, 2, money_cell(s['max_profit']))
    ws.cell(6, 1, "Max loss ($)"); ws.cell(6, 2, money_cell(s['max_loss']))
    ws.cell(7, 1, "Breakeven(s)")
    ws.cell(7, 2, ", ".join(f"{b:.2f}" for b in s['breakevens']) or "n/a")
    ws.cell(8, 1, "Capital ($)"); ws.cell(8, 2, money_cell(s['capital']))
    ws.cell(9, 1, "ROC at max profit")
    ws.cell(9, 2, (f"{s['roc_at_max_profit']*100:.1f}%"
                   if s['roc_at_max_profit'] is not None else "n/a"))
    if pop is not None:
        ws.cell(10, 1, "Prob. of profit (risk-neutral)"); ws.cell(10, 2, f"{pop*100:.0f}%")
    if strat.iv_percentile is not None:
        ws.cell(11, 1, "IV percentile"); ws.cell(11, 2, strat.iv_percentile)
        ws.cell(11, 3, _iv_bias_line(strat.iv_percentile))
    ws.cell(12, 1, "Model edge vs mids ($)"); ws.cell(12, 2, money_cell(s['model_edge']))
    ws.cell(12, 3, "~0 if contract IV and mid match; large = check inputs")

    title_row(14, "NET GREEKS (whole position)")
    greek_notes = {
        'delta': "directional bias (+ bullish)",
        'gamma': "convexity (+ you own acceleration)",
        'theta': "per DAY (+ you COLLECT decay = income)",
        'vega':  "per 1% IV (+ gain if IV rises)",
        'rho':   "per 1% rate move",
    }
    row = 15
    for k in ['delta', 'gamma', 'theta', 'vega', 'rho']:
        ws.cell(row, 1, k.capitalize()).font = bold
        ws.cell(row, 2, round(g[k], 3))
        ws.cell(row, 3, greek_notes[k])
        row += 1

    title_row(row + 1, "LEGS")
    row += 2
    headers = ["Type", "Strike", "Qty", "Mid", "Model", "IV typed", "IV from mid",
               "Delta $", "Theta $", "Vega $"]
    for h, col in zip(headers, range(1, len(headers) + 1)):
        ws.cell(row, col, h).font = bold
    row += 1
    for d in strat.per_leg_detail():
        leg = d['leg']
        ws.cell(row, 1, leg.kind)
        ws.cell(row, 2, leg.strike if leg.kind != 'stock' else "")
        ws.cell(row, 3, leg.quantity)
        ws.cell(row, 4, leg.premium)
        ws.cell(row, 5, (round(d['model_price'], 4) if d['model_price'] is not None else ""))
        ws.cell(row, 6, f"{leg.iv*100:.2f}%" if leg.kind != 'stock' else "")
        ws.cell(row, 7, (f"{d['implied_vol']*100:.2f}%" if d['implied_vol'] is not None else ""))
        ws.cell(row, 8, round(d['greeks']['delta'], 2))
        ws.cell(row, 9, round(d['greeks']['theta'], 3))
        ws.cell(row, 10, round(d['greeks']['vega'], 3))
        row += 1

    ws2 = wb.create_sheet("Payoff")
    ws2.cell(1, 1, "Stock Price").font = bold
    ws2.cell(1, 2, "P/L at expiration").font = bold
    ws2.cell(1, 3, "P/L T+0 (model)").font = bold
    prices, profits = strat.payoff_curve(points=60)
    t0 = [strat.mark_pnl(p) for p in prices]
    for i, (p, pf, now) in enumerate(zip(prices, profits, t0), start=2):
        ws2.cell(i, 1, round(float(p), 2))
        ws2.cell(i, 2, round(float(pf), 2))
        ws2.cell(i, 3, round(float(now), 2))
    chart = LineChart()
    chart.title = f"Payoff — {strat.name}"
    chart.x_axis.title = "Stock price"
    chart.y_axis.title = "P/L ($)"
    data = Reference(ws2, min_col=2, min_row=1, max_col=3, max_row=len(prices) + 1)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=len(prices) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws2.add_chart(chart, "E2")

    for col, w in zip("ABCDEFGHIJ", [28, 14, 14, 12, 12, 12, 14, 12, 12, 12]):
        ws.column_dimensions[col].width = w
    wb.save(path)


# =============================================================================
# WORKED EXAMPLE — run this file directly to see it in action.
# =============================================================================
# The "if __name__ == '__main__'" block only runs when you execute THIS file
# directly (python options_evaluator.py). It won't run if you import the
# functions elsewhere. Standard Python pattern for "demo / entry point."
#
# Day-to-day, use evaluate.py instead — that's the control panel with live
# chain numbers. This block is just a built-in demo of all three shapes.
# =============================================================================

if __name__ == "__main__":
    here = Path(__file__).resolve().parent

    # Illustrative numbers — replace via evaluate.py with a live chain.
    MRVL = 182.0
    RATE = 0.045
    DTE  = 45

    print("\n### THREE WAYS TO PLAY MRVL @ $182 - same view, different risk shapes ###\n")

    s1 = long_call(stock_price=MRVL, strike=190, iv=0.65, premium=9.50,
                   days_to_expiry=DTE, r=RATE, iv_percentile=72)
    print_summary(s1)

    s2 = cash_secured_put(stock_price=MRVL, put_strike=170, put_iv=0.68,
                          put_premium=6.20, days_to_expiry=DTE, r=RATE,
                          iv_percentile=72)
    print_summary(s2)

    s3 = covered_call(stock_price=MRVL, stock_entry=182, call_strike=200,
                      call_iv=0.63, call_premium=5.40, days_to_expiry=DTE,
                      r=RATE, iv_percentile=72)
    print_summary(s3)

    chart_path = here / "mrvl_csp_payoff.png"
    xlsx_path = here / "mrvl_strategy.xlsx"
    make_payoff_chart(s2, chart_path)
    export_excel(s2, xlsx_path)
    print(f"Saved: {chart_path.name}  and  {xlsx_path.name}")


# =============================================================================
# >>> HOW TO EXTEND THIS YOURSELF (your learning exercise) <<<
# Adding a new strategy is ALWAYS the same three steps:
#   1. Write a builder function (like covered_call) that assembles the right
#      OptionLegs with the correct quantity signs (+1 long, -1 short).
#   2. Hand those legs to Strategy(...).
#   3. Call print_summary / make_payoff_chart / export_excel on it.
# Try building a VERTICAL SPREAD next: long one call, short a higher-strike call
# (two legs, quantities +1 and -1). It's the exact structure we discussed for
# Marvell — and once you build it, you'll SEE in the payoff diagram how selling
# the upper leg caps your profit but slashes your cost and vega. That's the
# whole lesson, made visual. Good next step.
# =============================================================================
